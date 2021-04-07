import os
from openpyxl import load_workbook

class Environment:
    luaFiles: dict

    def __init__(self, *args):
        self.luaFiles = {}

        if args:
            for arg in args:
                if isinstance(arg, str):
                    if os.path.isfile(arg):
                        self.parseXlsx(arg)
                    elif os.path.isdir(arg):
                        self.parseDir(arg)

    def parseDir(self, dirPath):
        files = os.listdir(dirPath)
        for i in range(0, len(files)):
            path = os.path.join(dirPath, files[i])
            if os.path.isdir(path):
                self.parseDir(path)
            elif os.path.isfile(path):
                self.parseXlsx(path)


    def parseXlsx(self, value):
        _, filename = os.path.split(value)
        if filename[0] == '~':
            return

        if os.path.splitext(value)[-1] not in [".xlsx"]:
            return

        wb = load_workbook(filename=value, data_only=True)
        
        index = wb['INDEX']

        for rx in range(2, index.max_row + 1):
            if index.cell(row = rx, column=2).value.split(".")[-1] == "lua":
                sheetIndex = index.cell(row = rx, column=1).value
                self.generateLua(wb.worksheets[int(sheetIndex)], index.cell(row = rx, column=2).value)

    def generateLua(self, sheet, fileName):
        sb = []
        header = "local cls = {\n"
        ender = "\n}\nreturn cls"

        for i in range(3, sheet.max_row + 1):
            lineBulder = []
            for j in range(2, sheet.max_column + 1):
                if sheet.cell(row=2, column = j).value != "" and sheet.cell(row=2, column = j).value != None:
                    lineBulder.append(sheet.cell(row=2, column = j).value + "=\"" + sheet.cell(row=i, column = j).value + "\"")
            
            lineStr = sheet.cell(row=i, column = 1).value + " = { " + ','.join(lineBulder) + "}"

            sb.append(lineStr)
        
        luaStr = header + ",\n".join(sb) + ender
        
        self.luaFiles[fileName] = luaStr

            
            



